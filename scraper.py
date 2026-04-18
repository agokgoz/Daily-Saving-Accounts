"""
Daily Savings Account Interest Rate Scraper
============================================
Scrapes daily Welcome Rates for Turkish "günlük kazandıran hesaplar"
(daily savings accounts) from bank websites, compares them with yesterday's
rates stored in an Excel file, sends an HTML email notification if any rate
changes are detected, and appends today's rates to the Excel file.

Environment Variables Required (for email):
  SMTP_EMAIL    – sender Gmail address
  SMTP_PASSWORD – Gmail App Password
  TARGET_EMAIL  – recipient email address

Usage:
  python scraper.py
"""

import os
import re
import smtplib
import traceback
from datetime import date, datetime, timezone, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from html import escape as html_escape

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError

# ---------------------------------------------------------------------------
# Bank Configuration
# ---------------------------------------------------------------------------
# Each entry maps a bank name to:
#   url           – page to visit
#   welcome_sel   – CSS selector for the "hoş geldin / welcome rate" element
#   standard_sel  – CSS selector for the "standart / standard rate" element
#
# NOTE: CSS selectors are site-specific and must be verified / updated
# whenever a bank redesigns their page.  The selectors below were determined
# from the publicly available pages at the time of writing.  If a selector
# stops working, inspect the target page with your browser's DevTools and
# update the value here.
# ---------------------------------------------------------------------------

BANK_CONFIG: dict[str, dict] = {
    "ING (Turuncu Hesap)": {
        "url": "https://www.ing.com.tr/tr/bilgi-destek/e-turuncu-faiz-oranlari",
        "custom_scraper": "ing",
    },
    "Akbank (Serbest Hesap)": {
        "url": "https://www.akbank.com/mevduat-yatirim/mevduat/vadeli-mevduat-hesaplari/serbest-plus-hesap",
        "custom_scraper": "akbank",
    },
    "CEPTETEB (Marifetli Hesap)": {
        "url": "https://www.cepteteb.com.tr/hesaplama/marifetli-hesap-faiz-hesaplama",
        "custom_scraper": "teb",
    },
    "QNB (Kazandıran Günlük Hesap)": {
        "url": "https://www.qnb.com.tr/kazandiran-gunluk-hesap",
        "custom_scraper": "qnb",
    },
    "Enpara (Birikim Hesabı)": {
        "url": "https://www.enpara.com/hesaplar/birikim-hesabi#faiz-oranlari",
        "custom_scraper": "enpara",
    },
    "VakıfBank (ARI Hesabı)": {
        "url": "https://www.vakifbank.com.tr/tr/bireysel/hesaplar/vadeli-hesaplar/ari-hesabi",
        "custom_scraper": "vakifbank",
    },
    "Fibabanka (Kiraz Hesap)": {
        "url": "https://www.fibabanka.com.tr/mevduat/kiraz-hesap",
        "custom_scraper": "fibabanka",
    },
    "GetirFinans (Hesap)": {
        "url": "https://www.getirfinans.com/hesap/",
        "custom_scraper": "getirfinans",
    },
}

# Path of the Excel file that stores historical rates.
EXCEL_FILE = "historical_rates.xlsx"

# Default page-load / selector timeout in milliseconds.
PAGE_TIMEOUT_MS = 30_000


# ---------------------------------------------------------------------------
# Bank-Specific Scraping Functions
# ---------------------------------------------------------------------------
# Some banks require specialized scraping logic that uses text locators
# instead of brittle CSS selectors. These custom functions navigate the DOM
# using visible text labels to find rate values more reliably.
# ---------------------------------------------------------------------------

def clean_rate_text(text: str) -> str:
    """
    Clean a rate string by removing the '%' sign, trimming whitespace,
    and replacing Turkish decimal commas with dots.
    Returns the cleaned string (e.g., '53' or '23.5').
    """
    cleaned = text.strip().replace("%", "").replace(",", ".").strip()
    return cleaned


def parse_rate_float(text: str) -> float:
    """
    Parse a cleaned rate string into a float.
    Returns 0.0 if parsing fails.
    """
    print(f"    [DEBUG] Raw extracted text: '{text}'")
    try:
        return float(clean_rate_text(text))
    except (ValueError, AttributeError):
        return 0.0


def extract_rate_via_js(page, keyword: str, bank_name: str) -> float:
    """
    Extract a rate value using advanced JavaScript evaluation.
    Filters for visible elements, searches deepest nodes first, and climbs DOM.
    """
    try:
        page.wait_for_timeout(2000) # Give client-side frameworks a moment
        js_code = f"""
        () => {{
            const elements = Array.from(document.querySelectorAll('*'));
            
            // Find ALL elements containing the keyword that are actually VISIBLE
            const visibleTargets = elements.filter(el => {{
                const isVisible = el.offsetWidth > 0 && el.offsetHeight > 0;
                const text = el.innerText || el.textContent || "";
                return isVisible && text.toLowerCase().includes('{keyword}'.toLowerCase());
            }});
            
            if (visibleTargets.length === 0) return "";
            
            // Reverse the array to start with the deepest child elements first
            for (let target of visibleTargets.reverse()) {{
                let current = target;
                let matches = null;
                
                // Check the target itself and climb up to 7 parent levels
                for (let i = 0; i < 7; i++) {{
                    if (!current || current.tagName === 'BODY') break;
                    
                    const text = current.innerText || current.textContent || "";
                    matches = text.match(/%\\s?\\d+[.,]?\\d*|\\d+[.,]?\\d*\\s?%/g);
                    
                    if (matches) break;
                    current = current.parentElement;
                }}
                
                // If we found a percentage near this specific visible keyword, return it
                if (matches) {{
                    const numbers = matches.map(m => parseFloat(m.replace('%', '').replace(',', '.').trim()));
                    return Math.max(...numbers).toString();
                }}
            }}
            
            return "";
        }}
        """
        rate_text = page.evaluate(js_code)
        print(f"    [DEBUG] {bank_name} '{keyword}' raw JS extraction: '{rate_text}'")
        return parse_rate_float(rate_text)
    except Exception as exc:
        print(f"  [WARN] JS extraction failed for {bank_name} ({keyword}): {exc}")
        return 0.0


def get_ing_rates(page) -> dict[str, float]:
    return {
        "welcome_rate": extract_rate_via_js(page, "Hoş Geldin", "ING"),
    }


def get_akbank_rates(page) -> dict[str, float]:
    try:
        # Wait for the table that contains the 10.000 tier to load
        page.wait_for_selector("table:has-text('10.000')", timeout=10000)
        
        # Grab the text of the ENTIRE table, not just one row
        table = page.locator("table:has-text('10.000')").first
        table_text = table.inner_text()
        
        print(f"    [DEBUG] Akbank table text: '{table_text.replace(chr(10), ' ')}'")
        
        # Extract every valid percentage from the table
        matches = re.findall(r'%\s?\d+[.,]?\d*|\d+[.,]?\d*\s?%', table_text)
        
        if matches:
            # Strip the % signs and convert all matches to float
            numbers = [float(m.replace('%', '').replace(',', '.').strip()) for m in matches]
            
            # The 10.000 TL tier is the first column, so we want the FIRST percentage found
            return {"welcome_rate": numbers[0]}
            
    except Exception as exc:
        print(f"  [WARN] Table extraction failed for Akbank: {exc}")
        
    return {"welcome_rate": 0.0}

def get_qnb_rates(page) -> dict[str, float]:
    try:
        # Force Playwright to wait for the dynamic content container
        page.wait_for_selector('.table-wrap, .template-InterestRates, li:has-text("tanışma faizi")', timeout=10000)
        page.wait_for_timeout(2000)
    except Exception:
        pass # Proceed anyway if timeout occurs
        
    return {
        "welcome_rate": extract_rate_via_js(page, "Tanışma", "QNB"),
    }


def get_teb_rates(page) -> dict[str, float]:
    """
    Extract TEB Marifetli Hesap welcome rate using JavaScript evaluation.

    Returns:
        {"welcome_rate": float}
    """
    return {
        "welcome_rate": extract_rate_via_js(page, "Hoş Geldin", "TEB"),
    }


def get_enpara_rates(page) -> dict[str, float]:
    """
    Intercepts GetDailyInterestRates XHR calls (one per tier).
    "Evet" = Ayın Enparalısı rate. Returns the highest across all tiers.
    """
    evet_values = []

    def handle_response(response):
        if "GetDailyInterestRates" not in response.url:
            return
        if response.status != 200:
            return
        try:
            data = response.json()
            if not data.get("IsSucceded"):
                return
            for item in data.get("TransactionResult", []):
                if item.get("Key") == "Evet":
                    evet_values.append(float(item["Value"]))
        except Exception as e:
            print(f"    [DEBUG] Enpara XHR parse error: {e}")

    page.on("response", handle_response)

    # Land on homepage first to avoid bot detection on direct URL
    page.goto("https://www.enpara.com", wait_until="domcontentloaded", timeout=PAGE_TIMEOUT_MS)
    page.wait_for_timeout(2000)

    page.goto(
        "https://www.enpara.com/hesaplar/birikim-hesabi",
        wait_until="networkidle",
        timeout=PAGE_TIMEOUT_MS,
    )
    page.wait_for_timeout(2000)

    current_url = page.url
    print(f"    [DEBUG] Enpara landed on: {current_url}")
    if "birikim" not in current_url:
        print("    [DEBUG] Enpara: redirected, bot detection not bypassed")
        return {"welcome_rate": 0.0}

    # Scroll to trigger all tier XHR calls
    page.evaluate("document.getElementById('faiz-oranlari')?.scrollIntoView()")
    page.wait_for_timeout(2000)

    print(f"    [DEBUG] Enpara Evet values: {evet_values}")

    if evet_values:
        return {"welcome_rate": max(evet_values)}

    print("    [DEBUG] Enpara: no Evet values captured")
    return {"welcome_rate": 0.0}


def get_vakifbank_rates(page) -> dict[str, float]:
    """
    Bypass the dynamic calculator trap entirely and extract the maximum 
    percentage directly from VakıfBank's static interest rate table.
    """
    try:
        page.wait_for_selector("table", timeout=10000)
        
        # Target the table that contains the deposit terms (usually contains "Gün")
        table = page.locator("table:has-text('Gün')").first
        table_text = table.inner_text()
        
        print(f"    [DEBUG] VakıfBank table text: '{table_text.replace(chr(10), ' ')}'")
        
        # Strict Regex: Extract all valid percentages from the table
        matches = re.findall(r'%\s?\d+[.,]?\d*|\d+[.,]?\d*\s?%', table_text)
        
        if matches:
            # Strip the % signs and convert to float
            numbers = [float(m.replace('%', '').replace(',', '.').strip()) for m in matches]
            
            # Return the highest percentage found in the table
            return {"welcome_rate": max(numbers)}
            
    except Exception as exc:
        print(f"  [WARN] Table extraction failed for VakıfBank: {exc}")
        
    return {"welcome_rate": 0.0}


def get_fibabanka_rates(page) -> dict[str, float]:
    """
    Scrapes the Hoş Geldin rate from the #faiz-oranlari table.
    Cells with headers containing 'col4' hold the welcome rate column.
    The max across all tiers is returned (standard vs digital channels).
    """
    try:
        page.goto(
            "https://www.fibabanka.com.tr/mevduat/kiraz-hesap",
            wait_until="networkidle",
            timeout=PAGE_TIMEOUT_MS,
        )
        page.wait_for_timeout(2000)

        # Scroll to the rates section to ensure it renders
        page.evaluate("document.getElementById('faiz-oranlari')?.scrollIntoView()")
        page.wait_for_timeout(1500)

        # Wait for at least one col4 rate cell to appear in the DOM
        # ~= is the CSS word-match selector: matches 'col4' as an exact token
        # in the space-separated headers attribute, excluding col40, col45, etc.
        page.wait_for_selector("#faiz-oranlari td[headers~='col4']", timeout=10000)

        # Extract all welcome rate cells (col4 = Hoş Geldin columns)
        rates = page.evaluate("""
        () => {
            const cells = document.querySelectorAll("#faiz-oranlari td[headers~='col4']");
            const results = [];
            cells.forEach(cell => {
                const text = cell.innerText.replace('%', '').replace(',', '.').trim();
                const num = parseFloat(text);
                if (!isNaN(num) && num > 1 && num < 100) results.push(num);
            });
            return results;
        }
        """)

        print(f"    [DEBUG] Fibabanka col4 rates: {rates}")
        if rates:
            return {"welcome_rate": max(rates)}

    except Exception as exc:
        print(f"    [DEBUG] Fibabanka table scrape failed: {exc}")

    return {"welcome_rate": 0.0}


def get_getirfinans_rates(page) -> dict[str, float]:
    """
    Scrapes GetirFinans daily rate from the rate table.
    Targets <p> tags inside cells with 'text-right font-semibold' classes.
    The first cell is the standard rate (43%) — max would also work
    since the high-balance tier (30%) is lower.
    """
    try:
        page.goto(
            "https://www.getirfinans.com/hesap/",
            wait_until="networkidle",
            timeout=PAGE_TIMEOUT_MS,
        )
        page.wait_for_timeout(2000)

        # Rate cells are <p> tags inside divs with these exact Tailwind classes
        rate = page.evaluate("""
        () => {
            const cells = document.querySelectorAll('div[class*="text-right"][class*="font-semibold"] p');
            const rates = [];
            cells.forEach(cell => {
                const text = cell.innerText.replace('%', '').replace(',', '.').trim();
                const num = parseFloat(text);
                if (!isNaN(num) && num > 1 && num < 100) rates.push(num);
            });
            // Return the highest rate that is NOT the anomalous low-balance tier
            // (the 30% tier applies only to >5.4M TL balances — not the welcome rate)
            // Standard welcome rate (43%) appears in the majority of tiers
            const filtered = rates.filter(r => r > 35);
            return filtered.length > 0 ? Math.max(...filtered) : 0;
        }
        """)

        print(f"    [DEBUG] GetirFinans rate from table: {rate}")
        if rate:
            return {"welcome_rate": float(rate)}

    except Exception as exc:
        print(f"    [DEBUG] GetirFinans table scrape failed: {exc}")

    return {"welcome_rate": 0.0}


def scrape_all_banks() -> dict[str, dict[str, float]]:
    """
    Visit every bank URL with a single Playwright browser session and collect
    the Welcome Rate for each bank.

    Returns a nested dict:
      { bank_name: { "welcome_rate": float } }
    """
    # Map of custom scraper identifiers to functions
    CUSTOM_SCRAPERS = {
        "ing": get_ing_rates,
        "akbank": get_akbank_rates,
        "qnb": get_qnb_rates,
        "teb": get_teb_rates,
        "enpara": get_enpara_rates,
        "vakifbank": get_vakifbank_rates,
        "fibabanka": get_fibabanka_rates,
        "getirfinans": get_getirfinans_rates,
    }

    results: dict[str, dict[str, str]] = {}

    with sync_playwright() as pw:
        # Visible browser for debugging (headless=False)
        browser = pw.chromium.launch(headless=True)
        context = browser.new_context(
            # Mimic a regular desktop browser to avoid bot-detection.
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            locale="tr-TR",
        )
        page = context.new_page()

        for bank_name, config in BANK_CONFIG.items():
            print(f"Scraping: {bank_name}")
            welcome_rate = 0.0
            try:
                page.goto(config["url"], wait_until="domcontentloaded", timeout=PAGE_TIMEOUT_MS)
                
                # Wait for observation (to see Cloudflare checks, cookie banners, etc.)
                page.wait_for_timeout(3000)

                # Attempt to dismiss cookie banners
                try:
                    # Using a CSS selector list for common Turkish accept buttons
                    cookie_btn = page.locator(
                        "button:has-text('Kabul Et'), "
                        "button:has-text('Tümünü Kabul Et'), "
                        "button:has-text('Anladım'), "
                        "button:has-text('Tercihlerimi Kaydet'), "
                        "a:has-text('Anladım')"
                    ).first
                    
                    # Try to click it if it appears within 2 seconds
                    cookie_btn.click(timeout=2000)
                    # Wait a moment for the banner animation to slide away
                    page.wait_for_timeout(1000)
                    print("  [DEBUG] Cookie banner dismissed.")
                except Exception:
                    # If no banner is found or it times out, safely ignore and continue
                    pass

                # Execute the custom scraper function
                custom_scraper_id = config["custom_scraper"]
                custom_fn = CUSTOM_SCRAPERS[custom_scraper_id]
                rates = custom_fn(page)
                # Keep rates as floats for proper Excel number formatting
                welcome_rate = float(rates.get("welcome_rate", 0.0))

                print(f"  Welcome Rate : {welcome_rate if welcome_rate else '(not found)'}")
            except Exception as exc:
                print(f"  [ERROR] Failed to load page for {bank_name}: {exc}")

            results[bank_name] = {
                "welcome_rate": welcome_rate,
            }

        browser.close()

    return results


# ---------------------------------------------------------------------------
# Excel / Pandas helpers
# ---------------------------------------------------------------------------

def load_last_row(excel_file: str) -> dict[str, dict[str, float]]:
    """
    Load the last row from *excel_file* and return the same nested dict
    structure as :func:`scrape_all_banks`.  The Excel file is expected to have
    a "Date" index column and one column per bank:
      "<Bank Name> Welcome Rate".

    Returns an empty dict if the file does not exist or is empty.
    """
    if not os.path.exists(excel_file):
        return {}

    try:
        df = pd.read_excel(excel_file, index_col=0)
    except Exception as exc:
        print(f"[WARN] Could not read {excel_file}: {exc}")
        return {}

    if df.empty:
        return {}

    last = df.iloc[-1]
    previous: dict[str, dict[str, float]] = {}

    for bank_name in BANK_CONFIG:
        w_col = f"{bank_name} Welcome Rate"
        w_val = last.get(w_col)
        previous[bank_name] = {
            "welcome_rate": float(w_val) if pd.notna(w_val) else 0.0,
        }

    return previous


def has_entry_for_date(excel_file: str, target_date: date) -> bool:
    """
    Return True if *excel_file* already contains a row for *target_date*.
    """
    if not os.path.exists(excel_file):
        return False

    try:
        df = pd.read_excel(excel_file, index_col=0)
    except Exception as exc:
        print(f"[WARN] Could not read {excel_file} for date check: {exc}")
        return False

    if df.empty:
        return False

    try:
        normalized_index = pd.to_datetime(df.index).date
        return target_date in normalized_index
    except Exception as exc:
        print(f"[WARN] Could not parse dates in {excel_file}: {exc}")
        return False


def append_to_excel(excel_file: str, today: date, scraped: dict[str, dict[str, float]], scrape_time: datetime) -> None:
    """
    Upsert *scraped* rates for *today* into *excel_file*.
    Creates the file with the correct columns if it does not exist yet.
    If a row for *today* already exists, it is overwritten with latest values.
    Includes the scraping time and adjusts column widths for readability.
    """
    # Build a flat row dict keyed by column names.
    # Use Union type to allow both string (Time) and float (rates) values
    row: dict[str, str | float] = {}
    row["Time"] = scrape_time.strftime("%H:%M:%S")
    for bank_name, rates in scraped.items():
        # Rates are stored as floats for proper Excel number formatting
        row[f"{bank_name} Welcome Rate"] = rates["welcome_rate"]

    new_df = pd.DataFrame([row], index=[pd.Timestamp(today)])
    new_df.index.name = "Date"

    if os.path.exists(excel_file):
        try:
            existing_df = pd.read_excel(excel_file, index_col=0)
            existing_df.index = pd.to_datetime(existing_df.index).normalize()
            today_ts = pd.Timestamp(today)
            # Ensure Time column exists in existing data
            if "Time" not in existing_df.columns:
                existing_df.insert(0, "Time", "")

            # Ensure all new columns exist in existing data
            for col in new_df.columns:
                if col not in existing_df.columns:
                    existing_df[col] = pd.NA

            # Ensure new row has all existing columns
            new_df = new_df.reindex(columns=existing_df.columns)

            if today_ts in existing_df.index:
                # Overwrite today's row with latest reading
                existing_df.loc[today_ts, :] = new_df.iloc[0]
                combined = existing_df
            else:
                combined = pd.concat([existing_df, new_df])
        except Exception as exc:
            print(f"[WARN] Could not read existing file for append: {exc}. Overwriting.")
            combined = new_df
    else:
        combined = new_df

    # Ensure Time column is first
    cols = combined.columns.tolist()
    if "Time" in cols:
        cols.remove("Time")
        cols = ["Time"] + cols
        combined = combined[cols]

    combined.to_excel(excel_file)

    # Adjust column widths for readability
    _adjust_column_widths(excel_file)

    print(f"Saved rates to {excel_file} ({len(combined)} rows total).")


def _adjust_column_widths(excel_file: str) -> None:
    """
    Adjust Excel column widths based on content for better readability.
    Also formats the Date index column to display dates without time.
    """
    try:
        wb = load_workbook(excel_file)
        ws = wb.active

        for col_idx, col in enumerate(ws.columns, 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)

            for cell in col:
                try:
                    cell_value = str(cell.value) if cell.value is not None else ""
                    max_length = max(max_length, len(cell_value))
                except Exception:
                    pass

            # Add padding and set minimum width
            adjusted_width = max(max_length + 2, 10)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Apply date format (YYYY-MM-DD) to the Date index column (column A, rows 2+)
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=1)
            if cell.value is not None:
                cell.number_format = 'YYYY-MM-DD'

        wb.save(excel_file)
    except Exception as exc:
        print(f"[WARN] Could not adjust column widths: {exc}")


# ---------------------------------------------------------------------------
# Diffing
# ---------------------------------------------------------------------------

def find_changes(
    previous: dict[str, dict[str, float]],
    current: dict[str, dict[str, float]],
) -> list[dict]:
    """
    Compare *current* rates against *previous* rates.

    Returns a list of change dicts:
      {
        "bank":      str,
        "rate_type": "Welcome Rate",
        "old":       float,
        "new":       float,
      }

    Only non-zero new values that differ from the previous value are reported.
    If there is no previous data (first run) no changes are reported.
    """
    if not previous:
        print("No previous data found – skipping diff (first run).")
        return []

    changes: list[dict] = []
    for bank_name, rates in current.items():
        prev = previous.get(bank_name, {})
        old_val = prev.get("welcome_rate", 0.0)
        new_val = rates.get("welcome_rate", 0.0)
        # Only flag a change when we actually scraped a new (non-zero)
        # value that differs from what was stored.
        if new_val and new_val != old_val:
            changes.append({
                "bank": bank_name,
                "rate_type": "Welcome Rate",
                "old": old_val if old_val else "(no data)",
                "new": new_val,
            })
    return changes


# ---------------------------------------------------------------------------
# Email
# ---------------------------------------------------------------------------

def _build_source_websites_section() -> str:
    """
    Build an HTML section listing all source websites from which rate data is collected.
    This is included in emails for regulatory compliance purposes.
    """
    sources_html = "".join(
        f'<li style="margin-bottom:4px;">'
        f'<strong>{html_escape(bank_name)}:</strong> '
        f'<a href="{html_escape(url)}" style="color:#3498db;">{html_escape(url)}</a>'
        f'</li>'
        for bank_name, config in BANK_CONFIG.items()
        if (url := config.get("url", ""))
    )
    
    section = f"""
      <hr style="border:none;border-top:1px solid #ddd;margin:20px 0;">
      <div style="font-family:Arial,sans-serif;font-size:12px;color:#7f8c8d;">
        <p style="margin-bottom:8px;"><strong>📋 Data Sources</strong></p>
        <p style="margin-bottom:8px;">
          The interest rate information in this report was gathered from the following official bank websites:
        </p>
        <ul style="margin:0;padding-left:20px;">
          {sources_html}
        </ul>
      </div>
    """
    return section


def build_html_email(changes: list[dict], today: date) -> str:
    """
    Build an HTML email body that lists all rate changes with source URLs.
    If no changes, returns a "no changes detected" message. Both email variants
    include a data sources section at the bottom for regulatory compliance.
    """
    # Get the source websites section for regulatory compliance
    sources_section = _build_source_websites_section()
    
    if not changes:
        # No changes detected – send a simple notification
        html = f"""
        <html>
        <body>
          <h2 style="font-family:Arial,sans-serif;">
            📊 Daily Savings Rate Report – {today.strftime('%d %B %Y')}
          </h2>
          <p style="font-family:Arial,sans-serif;">
            ✅ <strong>No rate changes detected today.</strong>
          </p>
          <p style="font-family:Arial,sans-serif;">
            All monitored bank interest rates remain unchanged from yesterday.
          </p>
          <p style="font-family:Arial,sans-serif;color:#7f8c8d;font-size:12px;">
            The full historical rates ledger is attached.<br>
            This email was sent automatically by the Daily Savings Rate Scraper.
          </p>
          {sources_section}
        </body>
        </html>
        """
        return html

    # Build table for changes
    rows_html = ""
    for change in changes:
        source_url = BANK_CONFIG.get(change["bank"], {}).get("url", "")
        source_link = (
            f'<br><a href="{source_url}" style="font-size:11px;color:#3498db;">{source_url}</a>'
            if source_url else ""
        )
        rows_html += (
            f"<tr>"
            f"<td style='padding:8px;border:1px solid #ddd;'>{change['bank']}{source_link}</td>"
            f"<td style='padding:8px;border:1px solid #ddd;'>{change['rate_type']}</td>"
            f"<td style='padding:8px;border:1px solid #ddd;color:#e74c3c;'>{change['old']}</td>"
            f"<td style='padding:8px;border:1px solid #ddd;color:#27ae60;'>{change['new']}</td>"
            f"</tr>"
        )

    html = f"""
    <html>
    <body>
      <h2 style="font-family:Arial,sans-serif;">
        📊 Daily Savings Rate Changes – {today.strftime('%d %B %Y')}
      </h2>
      <p style="font-family:Arial,sans-serif;">
        The following interest rate changes were detected:
      </p>
      <table style="border-collapse:collapse;font-family:Arial,sans-serif;width:100%;">
        <thead>
          <tr style="background-color:#2c3e50;color:#fff;">
            <th style="padding:10px;border:1px solid #ddd;text-align:left;">Bank / Product</th>
            <th style="padding:10px;border:1px solid #ddd;text-align:left;">Rate Type</th>
            <th style="padding:10px;border:1px solid #ddd;text-align:left;">Previous Rate</th>
            <th style="padding:10px;border:1px solid #ddd;text-align:left;">New Rate</th>
          </tr>
        </thead>
        <tbody>
          {rows_html}
        </tbody>
      </table>
      <p style="font-family:Arial,sans-serif;color:#7f8c8d;font-size:12px;">
        The full historical rates ledger is attached.<br>
        This email was sent automatically by the Daily Savings Rate Scraper.
      </p>
      {sources_section}
    </body>
    </html>
    """
    return html


def send_email(changes: list[dict], today: date) -> None:
    """
    Send an HTML notification email listing changes (or no-changes notice),
    with the Excel ledger attached.

    Reads SMTP credentials from environment variables:
      SMTP_EMAIL    – Gmail sender address
      SMTP_PASSWORD – Gmail App Password (not the account password)
      TARGET_EMAIL  – recipient address
    """
    smtp_email = os.environ.get("SMTP_EMAIL", "")
    smtp_password = os.environ.get("SMTP_PASSWORD", "")
    target_email = os.environ.get("TARGET_EMAIL", "")

    if not all([smtp_email, smtp_password, target_email]):
        print("[WARN] Email credentials not set – skipping email notification.")
        return

    # Use different subject line based on whether changes were detected
    if changes:
        subject = f"[Rate Alert] Daily Savings Rate Changes – {today.strftime('%d %B %Y')}"
    else:
        subject = f"[Rate Report] No Changes – {today.strftime('%d %B %Y')}"

    html_body = build_html_email(changes, today)

    # Outer container must be 'mixed' to support both HTML body and attachment
    msg = MIMEMultipart("mixed")
    msg["Subject"] = subject
    msg["From"] = smtp_email
    msg["To"] = target_email

    # Wrap the HTML part in a 'related' container (best practice for HTML emails)
    html_part = MIMEMultipart("alternative")
    html_part.attach(MIMEText(html_body, "html", "utf-8"))
    msg.attach(html_part)

    # Attach the Excel file if it exists
    if os.path.exists(EXCEL_FILE):
        from email.mime.base import MIMEBase
        from email import encoders
        with open(EXCEL_FILE, "rb") as f:
            attachment = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            attachment.set_payload(f.read())
            encoders.encode_base64(attachment)
            attachment.add_header(
                "Content-Disposition",
                "attachment",
                filename=f"historical_rates_{today.strftime('%Y%m%d')}.xlsx",
            )
            msg.attach(attachment)
        print(f"  [DEBUG] Excel file attached: {EXCEL_FILE}")
    else:
        print(f"  [WARN] Excel file not found, skipping attachment: {EXCEL_FILE}")

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(smtp_email, smtp_password)
            server.sendmail(smtp_email, target_email, msg.as_string())
        print(f"Email sent to {target_email} with {len(changes)} change(s).")
    except Exception as exc:
        print(f"[ERROR] Failed to send email: {exc}")
        traceback.print_exc()


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    # Use UTC+3 timezone for Turkey
    tz_utc3 = timezone(timedelta(hours=3))
    today = datetime.now(tz_utc3).date()
    scrape_time = datetime.now(tz_utc3)
    print(f"=== Daily Savings Rate Scraper – {today} at {scrape_time.strftime('%H:%M:%S')} ===\n")
    is_first_run_today = not has_entry_for_date(EXCEL_FILE, today)

    # 1. Load yesterday's rates from the Excel ledger.
    print("Loading previous rates from Excel…")
    previous_rates = load_last_row(EXCEL_FILE)

    # 2. Scrape today's rates.
    print("\nScraping current rates…")
    current_rates = scrape_all_banks()

    # 3. Detect changes.
    print("\nComparing rates…")
    changes = find_changes(previous_rates, current_rates)
    if changes:
        print(f"  {len(changes)} change(s) detected:")
        for c in changes:
            print(f"    {c['bank']} | {c['rate_type']}: {c['old']} → {c['new']}")
    else:
        print("  No rate changes detected.")

    # 4. Upsert today's rates to the Excel ledger only if:
    #    - This is the first run of today, OR
    #    - Changes were detected in a later run
    if is_first_run_today:
        print("\nFirst run of today – recording daily snapshot…")
        append_to_excel(EXCEL_FILE, today, current_rates, scrape_time)
    elif changes:
        print("\nChanges detected – updating today's ledger row…")
        append_to_excel(EXCEL_FILE, today, current_rates, scrape_time)
    else:
        print("\nNo changes detected on later run – skipping Excel update.")

    # 5. Send email only if:
    #    - This is the first run of today (including no-change report), OR
    #    - Changes were detected
    if is_first_run_today or changes:
        print("\nSending email notification…")
        send_email(changes, today)
    else:
        print("\nNo changes detected on later run – skipping email notification.")

    print("\nDone.")


if __name__ == "__main__":
    main()
